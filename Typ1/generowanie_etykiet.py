import sys
import fpdf
import openpyxl


class PDF(fpdf.FPDF):
    def basic_table(self, rows):
        for label_row in rows:
            for label_col in label_row:
                self.cell(95, 27, label_col, 1, align="C")
            self.ln()


# Constant variables
COLUMN1 = ["Номер поставки", "Номер места", "Модель", "Количество (шт)", "ПРОДАВЕЦ", "ПОКУПАТЕЛЬ"]
COLUMN2 = ["Shipment No.", "Package No.", "Article", "Quantity (pcs)", "SELLER", "BUYER"]
CONST1 = "SOME RANDOM COMPANY NAME"
CONST2 = "SOME RANDOM COMPANY NAME"


excel_file_name = input("Nazwa pliku z danymi: ")
excel_file = f"{excel_file_name}.xlsx"

try:
    dataframe = openpyxl.load_workbook(excel_file)
except:
    print("Nie ma takiego pliku!")
    escape = input("Nacisnij 'Enter', aby wyjsc...")
    sys.exit()

dataframe1 = dataframe.active

data = [[] for i in range(dataframe1.max_row)]
max_len = 0
font_size = 12

for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        data[row].append(str(col[row].value).strip())
    test_row = f"{data[row][2]}_{data[row][3]}_{data[row][4]}"
    if len(test_row) > max_len:
        max_len = len(test_row)


data = [
    [[COLUMN1[0], COLUMN2[0], data[i][0]],
     [COLUMN1[1], COLUMN2[1], data[i][1]],
     [COLUMN1[2], COLUMN2[2], f"{data[i][2]}_{data[i][3]}_{data[i][4]}"],
     [COLUMN1[3], COLUMN2[3], data[i][5]],
     [COLUMN1[4], COLUMN2[4], CONST2],
     [COLUMN1[5], COLUMN2[5], CONST1]] for i in range(len(data))
]

if 36 <= max_len < 45:
    font_size = 10
elif 46 <= max_len < 55:
    font_size = 8
elif max_len >= 54:
    font_size = 7


pdf = PDF("L")
pdf.add_font('DejaVu', fname="./fonts/ttf/DejaVuSansCondensed.ttf", uni=True)
pdf.set_font('DejaVu', size=font_size)
for i in range(len(data)):
    pdf.add_page()
    pdf.basic_table(data[i])

output_name = input("Jak ma sie nazywac plik z etykietami? ")
pdf.output(f"{output_name}.pdf")
