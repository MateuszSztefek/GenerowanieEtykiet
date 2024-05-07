import sys
import openpyxl as openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak

excel_file_name = input("Nazwa pliku z danymi: ")
excel_file = f"{excel_file_name}.xlsx"

output_name = input("Jak ma sie nazywac plik z etykietami? ")
doc = SimpleDocTemplate(f"{output_name}.pdf", pagesize=landscape(letter))
elements = []

try:
    dataframe = openpyxl.load_workbook(excel_file)
except:
    print("Nie ma takiego pliku!")
    escape = input("Nacisnij 'Enter', aby wyjsc...")
    sys.exit()

dataframe1 = dataframe.active

data = [[] for i in range(dataframe1.max_row)]

for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        data[row].append(str(col[row].value).strip())

data = [
    [["SENDER"],
     ["SOME RANDOM COMPANY NAME\nSOME RANDOM ADDRESS\nSOME RANDOM ADDRESS"],
     ["HOMOGENEUS BOX"],
     ["MANUFACTURER", "ORDER NUMBER"],
     [data[i][0], data[i][1]],
     ["ORIGIN", "LIBELLE ARTICLE"],
     [data[i][2], data[i][3]],
     ["BAR CODE", f"{data[i][4].strip()[-6]}  {data[i][4].strip()[-5]}  {data[i][4].strip()[-4]}  {data[i][4].strip()[-3]}  {data[i][4].strip()[-2]}  {data[i][4].strip()[-1]}"],
     [data[i][4]],
     ["SIZE", "QTY"],
     [data[i][5], data[i][6]],
     ["COMPANY", "DELIVERY ADDRESS"],
     ["RANDOM", "SOME RANDOM ADDRESS\n SOME RANDOM ADDRESS"],
     ["BOX NUMBER", "TOTAL BOXES"],
     [data[i][7], data[i][8]]] for i in range(len(data))
]

for label in data:
    t = Table(label, colWidths=(3 * inch, 7 * inch), rowHeights=(
        0.35 * inch, 0.7 * inch, 0.35 * inch, 0.35 * inch, 0.35 * inch, 0.35 * inch, 0.35 * inch, 0.35 * inch,
        0.35 * inch,
        0.35 * inch, 0.35 * inch, 0.35 * inch, 0.5 * inch, 0.35 * inch, 0.35 * inch))

    t.setStyle(TableStyle([('SPAN', (1, 0), (0, 0)),
                           ('SPAN', (0, 1), (1, 1)),
                           ('SPAN', (0, 2), (1, 2)),
                           ('SPAN', (1, 8), (1, 7)),
                           ('ALIGN', (0, 0), (2, 15), 'CENTER'),
                           ('VALIGN', (0, 0), (1, 14), 'TOP'),
                           ('FONTSIZE', (0, 0), (2, 15), 16),
                           ('FONTSIZE', (1, 7), (-1, 8), 25),
                           ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                           ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                           ]))

    elements.append(t)
    elements.append(PageBreak())

doc.build(elements)
